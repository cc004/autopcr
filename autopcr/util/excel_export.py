from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from ..module.modulebase import ResultTable

# === 样式配置 ===
color1 = "E3F2FD"  # 蓝色浅底
color2 = "E8F5E9"  # 绿色浅底
border_style = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def get_depth(defs):
    max_depth = 0
    for h in defs:
        if isinstance(h, str):
            max_depth = max(max_depth, 1)
        else:
            key = next(iter(h))
            max_depth = max(max_depth, 1 + get_depth(h[key]))
    return max_depth


def count_leaves(defs):
    total = 0
    for h in defs:
        if isinstance(h, str):
            total += 1
        else:
            key = next(iter(h))
            total += count_leaves(h[key])
    return total


def build_header(ws, defs, row, col, max_depth, merges, colors):
    for i, h in enumerate(defs):
        if len(colors) < 2:
            color = colors[0]
        else:
            color = colors[i % 2]
        if isinstance(h, str):
            ws.cell(row=row, column=col, value=h)
            ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=col).border = border_style
            ws.cell(row=row, column=col).alignment = center_align

            if row < max_depth:
                ws.merge_cells(start_row=row, start_column=col, end_row=max_depth, end_column=col)
            col += 1
        else:
            key = next(iter(h))
            children = h[key]
            span = count_leaves(children)

            ws.cell(row=row, column=col, value=key)
            ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=col).border = border_style
            ws.cell(row=row, column=col).alignment = center_align

            if span > 1:
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)

            col = build_header(ws, children, row + 1, col, max_depth, merges, [color])
    return col


def data_cells(defs, row, colors):
    result = []
    for i, defn in enumerate(defs):
        if len(colors) < 2:
            color = colors[0]
        else:
            color = colors[i % 2]
        if isinstance(defn, str):
            result.append((row.get(defn, ""), color))
        else:
            key = next(iter(defn))
            children = defn[key]
            group_val = row.get(key, {})
            if not isinstance(group_val, dict):
                result.append((group_val, color))
            else:
                result.extend(data_cells(children, group_val, [color]))
    return result


async def export_excel(table: ResultTable) -> BytesIO:
    header = table.header
    data = table.data
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    max_depth = get_depth(header)
    colors = [color1, color2]

    build_header(ws, header, 1, 1, max_depth, [], colors)

    # 填充数据
    start_row = max_depth + 1
    for ri, row_data in enumerate(data, start=start_row):
        row_cells = data_cells(header, row_data, colors)
        for ci, (val, color) in enumerate(row_cells, start=1):
            ws.cell(row=ri, column=ci, value=val)
            ws.cell(row=ri, column=ci).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            ws.cell(row=ri, column=ci).border = border_style
            ws.cell(row=ri, column=ci).alignment = center_align

        # ===== 自动列宽调整 =====
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                # 按行分割，找最长那一行
                lines = str(cell.value).split("\n")
                longest_line = max(len(line) for line in lines)
                max_length = max(max_length, longest_line)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 3

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
